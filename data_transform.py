import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# Define the Excel file path
excel_file_path = 'tasks.xlsx'

# Extract task data
task_label = "Task Label"  # Replace with the actual task label
task_name = "Task Name"    # Replace with the actual task name
start_time_str = "2023:10:08:18:47:20"  # Replace with the actual start time in the format YY:MM:DD:HH:MM:SS
end_time_str = "2023:10:08:19:47:20"    # Replace with the actual end time in the format YY:MM:DD:HH:MM:SS
duration_str = "01:00:00"  # Replace with the actual duration in the format HH:MM:SS

# Convert start and end times to datetime objects
start_time = datetime.strptime(start_time_str, '%Y:%m:%d:%H:%M:%S')
end_time = datetime.strptime(end_time_str, '%Y:%m:%d:%H:%M:%S')

# Calculate duration in seconds
duration_parts = duration_str.split(':')
duration_seconds = int(duration_parts[0]) * 3600 + int(duration_parts[1]) * 60 + int(duration_parts[2])

# Load the Excel file
try:
    workbook = load_workbook(excel_file_path)
except FileNotFoundError:
    # If the file doesn't exist, create a new workbook
    workbook = pd.ExcelWriter(excel_file_path, engine='openpyxl') 
    workbook.save()
    workbook = load_workbook(excel_file_path)

# Create a Pandas DataFrame with the task data
data = {
    'Task Label': [task_label],
    'Task Name': [task_name],
    'Start DateTime': [start_time],
    'End DateTime': [end_time],
    'Duration (seconds)': [duration_seconds]
}

df = pd.DataFrame(data)

# Append the DataFrame to the Excel file
with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    writer.book = workbook
    writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
    df.to_excel(writer, sheet_name='Task Data', index=False, header=False)

print("Task data appended to Excel file.")
